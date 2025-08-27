# -*- coding: utf-8 -*-
import sys, os, math, json, calendar
from datetime import datetime
from PySide6 import QtWidgets, QtGui, QtCore
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ASSETS = os.path.join(os.path.dirname(__file__), "..", "assets")
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
CONFIG_PATH = os.path.join(DATA_DIR, "config.json")


def load_config():
    default = {
        "neon": False,
        "header_font": "Arial",
        "text_font": "Arial",
        "save_path": DATA_DIR,
    }
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            save_path = data.get("save_path")
            if save_path:
                if not os.path.isabs(save_path):
                    save_path = os.path.abspath(
                        os.path.join(os.path.dirname(CONFIG_PATH), save_path)
                    )
                data["save_path"] = save_path
            default.update({k: v for k, v in data.items() if v is not None})
        except Exception:
            pass
    else:
        try:
            os.makedirs(os.path.dirname(CONFIG_PATH), exist_ok=True)
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(default, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    return default


CONFIG = load_config()
BASE_SAVE_PATH = os.path.abspath(CONFIG.get("save_path", DATA_DIR))
STATS_DIR = os.path.join(BASE_SAVE_PATH, "stats")
YEAR_DIR = os.path.join(BASE_SAVE_PATH, "year")
RELEASE_DIR = os.path.join(BASE_SAVE_PATH, "release")
TOP_DIR = os.path.join(BASE_SAVE_PATH, "top")
EXCEL_PATH = os.path.join(ASSETS, "пример блоков.xlsx")
SHEET_NAME = "ЦЕНТРАЛЬНАЯ РАБОЧАЯ ОБЛАСТЬ"
ICON_TOGGLE = os.path.join(ASSETS, "gpt_icon.png")
ICON_VYKL = os.path.join(ASSETS, "ic_vykl.png")
ICON_TM   = os.path.join(ASSETS, "ic_tm.png")
ICON_TQ   = os.path.join(ASSETS, "ic_tq.png")
ICON_TP   = os.path.join(ASSETS, "ic_tp.png")
ICON_TG   = os.path.join(ASSETS, "ic_tg.png")

RU_MONTHS = ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

# роли для хранения структуры работ и отметки "18+"
WORK_ROLE  = QtCore.Qt.UserRole
ADULT_ROLE = QtCore.Qt.UserRole + 1

def excel_col_width_to_pixels(width):
    if width is None: return 64
    return max(20, int((width * 7) + 5))

def excel_row_height_to_pixels(height_points):
    if height_points is None: return 20
    return max(16, int(round(height_points * (96.0/72.0))))

def qt_align(h, v):
    ah = {"left": QtCore.Qt.AlignLeft, "center": QtCore.Qt.AlignHCenter, "right": QtCore.Qt.AlignRight, "justify": QtCore.Qt.AlignJustify, None: QtCore.Qt.AlignLeft}.get(h, QtCore.Qt.AlignLeft)
    av = {"top": QtCore.Qt.AlignTop, "center": QtCore.Qt.AlignVCenter, "bottom": QtCore.Qt.AlignBottom, None: QtCore.Qt.AlignVCenter}.get(v, QtCore.Qt.AlignVCenter)
    return ah | av

class ExcelDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, borders_map, rows, cols, parent=None):
        super().__init__(parent); self.borders_map=borders_map; self.rows=rows; self.cols=cols
    def paint(self, painter, option, index):
        super().paint(painter, option, index)

        # borders from Excel
        r, c = index.row(), index.column()
        sides = self.borders_map.get((r, c))
        if sides:
            pen = QtGui.QPen(QtGui.QColor(0, 0, 0))
            pen.setWidth(1)
            painter.save()
            painter.setPen(pen)
            rect = option.rect.adjusted(0, 0, -1, -1)
            if sides.get("top"):
                painter.drawLine(rect.topLeft(), rect.topRight())
            if sides.get("left"):
                painter.drawLine(rect.topLeft(), rect.bottomLeft())
            if c == self.cols - 1 and sides.get("right"):
                painter.drawLine(rect.topRight(), rect.bottomRight())
            if r == self.rows - 1 and sides.get("bottom"):
                painter.drawLine(rect.bottomLeft(), rect.bottomRight())
            painter.restore()

        # mark 18+ cells
        if index.data(ADULT_ROLE):
            painter.save()
            painter.setPen(QtCore.Qt.NoPen)
            painter.setBrush(QtGui.QColor(200, 0, 0))
            rect = option.rect
            tri = QtGui.QPolygon(
                [rect.topRight(), rect.topRight() + QtCore.QPoint(-10, 0), rect.topRight() + QtCore.QPoint(0, 10)]
            )
            painter.drawPolygon(tri)
            painter.restore()

class WorkEditDialog(QtWidgets.QDialog):
    """Простое редактирование списка работ в ячейке."""

    def __init__(self, works, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Работы")
        self.resize(300, 200)
        self.works = works[:]

        self.lay = QtWidgets.QVBoxLayout(self)
        self.rows = []

        for w in self.works:
            self._add_row(w)

        btn_add = QtWidgets.QPushButton("Добавить", self)
        btn_add.clicked.connect(lambda: self._add_row({"plan": "", "done": False, "is_adult": False}))
        self.lay.addWidget(btn_add)

        box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, self)
        box.accepted.connect(self.accept)
        box.rejected.connect(self.reject)
        self.lay.addWidget(box)

    def _add_row(self, data):
        row_w = QtWidgets.QWidget(self)
        hl = QtWidgets.QHBoxLayout(row_w)
        hl.setContentsMargins(0, 0, 0, 0)
        cb_done = QtWidgets.QCheckBox("", row_w)
        cb_done.setChecked(data.get("done", False))
        hl.addWidget(cb_done)
        le = QtWidgets.QLineEdit(data.get("plan", ""), row_w)
        hl.addWidget(le, 1)
        cb_adult = QtWidgets.QCheckBox("18+", row_w)
        cb_adult.setChecked(data.get("is_adult", False))
        hl.addWidget(cb_adult)
        self.lay.addWidget(row_w)
        self.rows.append((cb_done, le, cb_adult))

    def get_works(self):
        works = []
        for cb_done, le, cb_adult in self.rows:
            txt = le.text().strip()
            if not txt:
                continue
            works.append({"plan": txt, "done": cb_done.isChecked(), "is_adult": cb_adult.isChecked()})
        return works


class ReleaseDialog(QtWidgets.QDialog):
    """Диалог для управления выкладками."""

    def __init__(self, year, month, parent=None):
        super().__init__(parent)
        self.year = year
        self.month = month
        self.setWindowTitle("Выкладка")
        self.resize(500, 300)

        lay = QtWidgets.QVBoxLayout(self)
        self.table = QtWidgets.QTableWidget(0, 4, self)
        self.table.setHorizontalHeaderLabels(["Дата", "Работа", "Глав", "Время"])
        self.table.horizontalHeader().setStretchLastSection(True)
        lay.addWidget(self.table)

        btn_row = QtWidgets.QHBoxLayout()
        btn_add = QtWidgets.QPushButton("Добавить", self)
        btn_add.clicked.connect(self.add_row)
        btn_del = QtWidgets.QPushButton("Удалить", self)
        btn_del.clicked.connect(self.remove_row)
        btn_row.addWidget(btn_add)
        btn_row.addWidget(btn_del)
        btn_row.addStretch(1)
        lay.addLayout(btn_row)

        box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Close, self)
        box.accepted.connect(self.save)
        box.rejected.connect(self.reject)
        lay.addWidget(box)

        self.load()

    def file_path(self):
        os.makedirs(RELEASE_DIR, exist_ok=True)
        return os.path.join(RELEASE_DIR, f"{self.year:04d}-{self.month:02d}.json")

    def add_row(self, data=None):
        if data is None:
            data = {"date": "", "work": "", "chapters": "", "time": ""}
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QtWidgets.QTableWidgetItem(str(data.get("date", ""))))
        self.table.setItem(row, 1, QtWidgets.QTableWidgetItem(data.get("work", "")))
        self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(data.get("chapters", ""))))
        self.table.setItem(row, 3, QtWidgets.QTableWidgetItem(data.get("time", "")))

    def remove_row(self):
        row = self.table.currentRow()
        if row >= 0:
            self.table.removeRow(row)

    def load(self):
        path = self.file_path()
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            for rec in data:
                self.add_row(rec)

    def save(self):
        data = []
        for r in range(self.table.rowCount()):
            date = self.table.item(r, 0)
            work = self.table.item(r, 1)
            chapters = self.table.item(r, 2)
            time = self.table.item(r, 3)
            if any(it and it.text().strip() for it in [date, work, chapters, time]):
                data.append({
                    "date": date.text().strip() if date else "",
                    "work": work.text().strip() if work else "",
                    "chapters": chapters.text().strip() if chapters else "",
                    "time": time.text().strip() if time else "",
                })
        with open(self.file_path(), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.accept()

class YearStatsDialog(QtWidgets.QDialog):
    """Годовая статистика: месяцы × показатели с колонкой "Итого за год"."""

    INDICATORS = [
        "Работ", "Завершённых", "Онгоингов", "Глав", "Знаков",
        "Просмотров", "Профит", "Реклама (РК)", "Лайков", "Спасибо",
        "Комиссия", "Затраты на софт", "Чистыми",
    ]

    def __init__(self, year, parent=None):
        super().__init__(parent)
        self.year = year
        self.setWindowTitle("Статистика")
        self.resize(900, 400)

        lay = QtWidgets.QVBoxLayout(self)
        top = QtWidgets.QHBoxLayout()
        top.addWidget(QtWidgets.QLabel("Год:"))
        self.spin_year = QtWidgets.QSpinBox(self)
        self.spin_year.setRange(2000, 2100)
        self.spin_year.setValue(year)
        self.spin_year.valueChanged.connect(self._year_changed)
        top.addWidget(self.spin_year)
        top.addStretch(1)
        self.btn_save = QtWidgets.QPushButton("Сохранить", self)
        self.btn_save.clicked.connect(self.save)
        top.addWidget(self.btn_save)
        lay.addLayout(top)

        cols = len(RU_MONTHS) + 1
        self.table = QtWidgets.QTableWidget(len(self.INDICATORS), cols, self)
        self.table.setHorizontalHeaderLabels(RU_MONTHS + ["Итого за год"])
        self.table.setVerticalHeaderLabels(self.INDICATORS)
        lay.addWidget(self.table, 1)

        # prepare items
        for r, name in enumerate(self.INDICATORS):
            for c in range(cols):
                it = QtWidgets.QTableWidgetItem("0")
                if name not in ("Комиссия", "Затраты на софт") or c == cols - 1:
                    it.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.table.setItem(r, c, it)

        self.table.itemChanged.connect(self._item_changed)

        self._loading = False
        self._commissions = {str(m): 0.0 for m in range(1, 13)}
        self._software = {str(m): 0.0 for m in range(1, 13)}
        self.load(year)

    def _year_changed(self, val):
        self.load(val)

    # --- data handling -------------------------------------------------
    def load(self, year):
        self._loading = True
        self.year = year
        self.spin_year.setValue(year)

        # load manual values
        self._commissions = {str(m): 0.0 for m in range(1, 13)}
        self._software = {str(m): 0.0 for m in range(1, 13)}
        path = os.path.join(YEAR_DIR, f"{year}.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            self._commissions.update({str(k): float(v) for k, v in data.get("commission", {}).items()})
            self._software.update({str(k): float(v) for k, v in data.get("software", {}).items()})

        # fill table with monthly values
        for m in range(1, 13):
            stats = self._calc_month_stats(year, m)
            for ind, val in stats.items():
                row = self.INDICATORS.index(ind)
                self.table.item(row, m - 1).setText(str(val))
            self.table.item(self.INDICATORS.index("Комиссия"), m - 1).setText(str(self._commissions[str(m)]))
            self.table.item(self.INDICATORS.index("Затраты на софт"), m - 1).setText(str(self._software[str(m)]))

        self._recalculate()
        self._loading = False

    def save(self):
        os.makedirs(YEAR_DIR, exist_ok=True)
        path = os.path.join(YEAR_DIR, f"{self.year}.json")
        data = {
            "commission": self._commissions,
            "software": self._software,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.accept()

    # --- helpers -------------------------------------------------------
    def _calc_month_stats(self, year, month):
        res = {k: 0 for k in self.INDICATORS if k not in ("Комиссия", "Затраты на софт", "Чистыми")}
        path = os.path.join(STATS_DIR, f"{year}.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            month_data = data.get(str(month), [])
            res["Работ"] = len(month_data)
            for rec in month_data:
                status = (rec.get("status", "") or "").lower()
                if "заверш" in status:
                    res["Завершённых"] += 1
                elif "онго" in status:
                    res["Онгоингов"] += 1
                res["Глав"] += int(rec.get("chapters", 0) or 0)
                res["Знаков"] += int(rec.get("chars", 0) or 0)
                res["Просмотров"] += int(rec.get("views", 0) or 0)
                res["Профит"] += float(rec.get("profit", 0) or 0)
                res["Реклама (РК)"] += float(rec.get("ads", 0) or 0)
                res["Лайков"] += int(rec.get("likes", 0) or 0)
                res["Спасибо"] += int(rec.get("thanks", 0) or 0)
        return res

    def _item_changed(self, item):
        if self._loading:
            return
        row = item.row()
        col = item.column()
        ind = self.INDICATORS[row]
        if ind == "Комиссия":
            try:
                self._commissions[str(col + 1)] = float(item.text())
            except ValueError:
                self._commissions[str(col + 1)] = 0.0
        elif ind == "Затраты на софт":
            try:
                self._software[str(col + 1)] = float(item.text())
            except ValueError:
                self._software[str(col + 1)] = 0.0
        self._recalculate()

    def _recalculate(self):
        row_profit = self.INDICATORS.index("Профит")
        row_comm = self.INDICATORS.index("Комиссия")
        row_soft = self.INDICATORS.index("Затраты на софт")
        row_net = self.INDICATORS.index("Чистыми")
        cols = len(RU_MONTHS)

        for c in range(cols):
            try:
                profit = float(self.table.item(row_profit, c).text())
            except ValueError:
                profit = 0.0
            comm = self._commissions[str(c + 1)]
            soft = self._software[str(c + 1)]
            net = profit - comm - soft
            it = self.table.item(row_net, c)
            it.setText(str(round(net, 2)))

        # totals
        for r in range(len(self.INDICATORS)):
            total = 0.0
            for c in range(cols):
                try:
                    total += float(self.table.item(r, c).text())
                except ValueError:
                    pass
            self.table.item(r, cols).setText(str(round(total, 2)))

class TopDialog(QtWidgets.QDialog):
    """Агрегирование и сохранение топов за период."""

    SORT_OPTIONS = [
        ("Профит", "profit"),
        ("Завершённость", "done"),
        ("Глав", "chapters"),
        ("Знаков", "chars"),
        ("Просмотров", "views"),
        ("Реклама (РК)", "ads"),
        ("Лайков", "likes"),
        ("Спасибо", "thanks"),
    ]

    def __init__(self, year, mode="month", period=1, parent=None):
        super().__init__(parent)
        self.year = year
        self.mode = mode
        self.period = period
        self.results = []
        self.setWindowTitle("Топ")
        self.resize(700, 400)

        lay = QtWidgets.QVBoxLayout(self)
        top = QtWidgets.QHBoxLayout()
        top.addWidget(QtWidgets.QLabel("Год:"))
        self.spin_year = QtWidgets.QSpinBox(self)
        self.spin_year.setRange(2000, 2100)
        self.spin_year.setValue(year)
        top.addWidget(self.spin_year)

        self.combo_period = None
        if mode != "year":
            self.combo_period = QtWidgets.QComboBox(self)
            if mode == "month":
                top.addWidget(QtWidgets.QLabel("Месяц:"))
                for i, m in enumerate(RU_MONTHS, 1):
                    self.combo_period.addItem(m, i)
                self.combo_period.setCurrentIndex(period - 1)
            elif mode == "quarter":
                top.addWidget(QtWidgets.QLabel("Квартал:"))
                for i in range(1, 5):
                    self.combo_period.addItem(str(i), i)
                self.combo_period.setCurrentIndex(period - 1)
            elif mode == "half":
                top.addWidget(QtWidgets.QLabel("Полугодие:"))
                for i in range(1, 3):
                    self.combo_period.addItem(str(i), i)
                self.combo_period.setCurrentIndex(period - 1)
            top.addWidget(self.combo_period)

        top.addWidget(QtWidgets.QLabel("Сортировка:"))
        self.combo_sort = QtWidgets.QComboBox(self)
        for label, key in self.SORT_OPTIONS:
            self.combo_sort.addItem(label, key)
        top.addWidget(self.combo_sort)

        self.btn_calc = QtWidgets.QPushButton("Сформировать", self)
        self.btn_calc.clicked.connect(self.calculate)
        top.addWidget(self.btn_calc)
        lay.addLayout(top)

        headers = [
            "Работа",
            "Статус",
            "Глав",
            "Знаков",
            "Просмотров",
            "Профит",
            "Реклама",
            "Лайков",
            "Спасибо",
        ]
        self.table = QtWidgets.QTableWidget(0, len(headers), self)
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setStretchLastSection(True)
        lay.addWidget(self.table, 1)

        box = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Close, self
        )
        box.accepted.connect(self.save)
        box.rejected.connect(self.reject)
        lay.addWidget(box)

    # --- helpers -------------------------------------------------------
    def _months_for_period(self):
        if self.mode == "month" and self.combo_period:
            return [self.combo_period.currentData()]
        if self.mode == "quarter" and self.combo_period:
            q = self.combo_period.currentData()
            start = (q - 1) * 3 + 1
            return list(range(start, start + 3))
        if self.mode == "half" and self.combo_period:
            h = self.combo_period.currentData()
            start = (h - 1) * 6 + 1
            return list(range(start, start + 6))
        return list(range(1, 13))

    def calculate(self):
        year = self.spin_year.value()
        months = self._months_for_period()
        sort_key = self.combo_sort.currentData()
        path = os.path.join(STATS_DIR, f"{year}.json")
        totals = {}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            for m in months:
                for rec in data.get(str(m), []):
                    work = rec.get("work", "")
                    t = totals.setdefault(
                        work,
                        {
                            "status": "",
                            "chapters": 0,
                            "chars": 0,
                            "views": 0,
                            "profit": 0.0,
                            "ads": 0.0,
                            "likes": 0,
                            "thanks": 0,
                            "done": 0,
                        },
                    )
                    t["status"] = rec.get("status", t["status"])
                    t["chapters"] += int(rec.get("chapters", 0) or 0)
                    t["chars"] += int(rec.get("chars", 0) or 0)
                    t["views"] += int(rec.get("views", 0) or 0)
                    t["profit"] += float(rec.get("profit", 0) or 0)
                    t["ads"] += float(rec.get("ads", 0) or 0)
                    t["likes"] += int(rec.get("likes", 0) or 0)
                    t["thanks"] += int(rec.get("thanks", 0) or 0)
                    status = (rec.get("status", "") or "").lower()
                    if "заверш" in status:
                        t["done"] += 1

        results = sorted(totals.items(), key=lambda kv: kv[1].get(sort_key, 0), reverse=True)
        self.results = results
        self.table.setRowCount(0)
        for work, vals in results:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QtWidgets.QTableWidgetItem(work))
            self.table.setItem(row, 1, QtWidgets.QTableWidgetItem(vals.get("status", "")))
            self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(vals["chapters"])))
            self.table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(vals["chars"])))
            self.table.setItem(row, 4, QtWidgets.QTableWidgetItem(str(vals["views"])))
            self.table.setItem(row, 5, QtWidgets.QTableWidgetItem(str(round(vals["profit"], 2))))
            self.table.setItem(row, 6, QtWidgets.QTableWidgetItem(str(round(vals["ads"], 2))))
            self.table.setItem(row, 7, QtWidgets.QTableWidgetItem(str(vals["likes"])))
            self.table.setItem(row, 8, QtWidgets.QTableWidgetItem(str(vals["thanks"])))

    def _period_key(self):
        if self.mode == "month" and self.combo_period:
            return f"M{self.combo_period.currentData():02d}"
        if self.mode == "quarter" and self.combo_period:
            return f"Q{self.combo_period.currentData()}"
        if self.mode == "half" and self.combo_period:
            return f"H{self.combo_period.currentData()}"
        return "Y"

    def save(self):
        if not self.results:
            self.calculate()
        year = self.spin_year.value()
        os.makedirs(TOP_DIR, exist_ok=True)
        path = os.path.join(TOP_DIR, f"{year}.json")
        data = {}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        key = self._period_key()
        data[key] = {
            "sort": self.combo_sort.currentData(),
            "results": [{"work": w, **vals} for w, vals in self.results],
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.accept()

class ExcelCalendarTable(QtWidgets.QTableWidget):
    """Рендер Excel + управление днями месяца и «хвостом» следующего/предыдущего месяца."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        self.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.setShowGrid(False); self.verticalHeader().setVisible(False); self.horizontalHeader().setVisible(False)
        self.setWordWrap(True); self.setEditTriggers(QtWidgets.QAbstractItemView.AllEditTriggers)

        now = datetime.now()
        self.year = now.year
        self.month = now.month

        self.base_col_widths=[]; self.base_row_heights=[]
        self.week_blocks=[]  # list of (numbers_row, weekdays_row, content_start_row) per неделю
        self.day_cols=(None,None)  # (start, end) индексы 7 колонок дней

        self.load_excel(EXCEL_PATH, SHEET_NAME)
        self.apply_month_numbers()

        # редактирование содержимого ячеек (список работ)
        self.itemDoubleClicked.connect(self.edit_cell)

    def _works_to_text(self, works):
        lines = []
        for w in works:
            prefix = "[x]" if w.get("done") else "[ ]"
            txt = f"{prefix} {w.get('plan','')}"
            if w.get("is_adult"):
                txt += " (18+)"
            lines.append(txt)
        return "\n".join(lines)

    def edit_cell(self, item):
        works = item.data(WORK_ROLE) or []
        dlg = WorkEditDialog(works, self)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            works = dlg.get_works()
            item.setData(WORK_ROLE, works)
            item.setText(self._works_to_text(works))
            item.setData(ADULT_ROLE, any(w.get("is_adult") for w in works))

    # ---------- Persistence ----------
    def month_key(self, year=None, month=None):
        y = year or self.year; m = month or self.month
        return f"{y:04d}-{m:02d}"

    def save_current_month(self):
        path = os.path.join(DATA_DIR, self.month_key()+".json")
        data = {"rows": self.rowCount(), "cols": self.columnCount(), "grid": [], "works": {}}
        for r in range(self.rowCount()):
            row = []
            for c in range(self.columnCount()):
                it = self.item(r, c)
                row.append("" if it is None else it.text())
                if it:
                    works = it.data(WORK_ROLE) or []
                    if works:
                        data["works"][f"{r},{c}"] = works
            data["grid"].append(row)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)

    def load_month_data(self, year, month):
        path = os.path.join(DATA_DIR, f"{year:04d}-{month:02d}.json")
        if not os.path.exists(path): return False
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        works_map = data.get("works", {})
        for r, row in enumerate(data.get("grid", [])):
            for c, txt in enumerate(row):
                it = self.item(r, c) or QtWidgets.QTableWidgetItem()
                works = works_map.get(f"{r},{c}", [])
                it.setData(WORK_ROLE, works)
                it.setData(ADULT_ROLE, any(w.get("is_adult") for w in works))
                if works:
                    it.setText(self._works_to_text(works))
                else:
                    it.setText(txt)
                self.setItem(r, c, it)
        return True

    # ---------- Excel load ----------
    def load_excel(self, path, sheet_name):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        wb = load_workbook(path, data_only=True); ws = wb[sheet_name]
        min_row=min_col=max_row=max_col=None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.coordinate in ws.merged_cells:
                    r,c=cell.row,cell.column
                    min_row = r if min_row is None else min(min_row,r)
                    min_col = c if min_col is None else min(min_col,c)
                    max_row = r if max_row is None else max(max_row,r)
                    max_col = c if max_col is None else max(max_col,c)
        if min_row is None: min_row=min_col=max_row=max_col=1
        rows=max_row-min_row+1; cols=max_col-min_col+1
        self.setRowCount(rows); self.setColumnCount(cols)

        borders={}
        for r in range(min_row, max_row+1):
            for c in range(min_col, max_col+1):
                b = ws.cell(r,c).border
                borders[(r-min_row,c-min_col)] = {"left":bool(b.left and b.left.style), "right":bool(b.right and b.right.style), "top":bool(b.top and b.top.style), "bottom":bool(b.bottom and b.bottom.style)}
        self.setItemDelegate(ExcelDelegate(borders, rows, cols, self))

        # sizes
        self.base_col_widths=[excel_col_width_to_pixels(ws.column_dimensions.get(get_column_letter(c)).width if ws.column_dimensions.get(get_column_letter(c)) else None) for c in range(min_col, max_col+1)]
        self.base_row_heights=[excel_row_height_to_pixels(ws.row_dimensions.get(r).height if ws.row_dimensions.get(r) else None) for r in range(min_row, max_row+1)]
        for i,w in enumerate(self.base_col_widths): self.setColumnWidth(i,w)
        for i,h in enumerate(self.base_row_heights): self.setRowHeight(i,h)

        # content + record weekday rows
        default_font_family="Calibri"; default_font_size=11
        wk_names = ["ПН","ВТ","СР","ЧТ","ПТ","СБ","ВС"]
        self.week_blocks.clear()
        self.day_cols=(None,None)

        def is_weekdays_row(texts):
            return texts == wk_names

        for r in range(min_row, max_row+1):
            # set items
            for c in range(min_col, max_col+1):
                cell=ws.cell(r,c)
                it = QtWidgets.QTableWidgetItem("" if cell.value is None else str(cell.value))
                f=cell.font; qf=QtGui.QFont(f.name or default_font_family, int(f.sz or default_font_size)); qf.setBold(bool(f.b)); qf.setItalic(bool(f.i)); it.setFont(qf)
                al=cell.alignment; it.setTextAlignment(qt_align(al.horizontal, al.vertical))
                it.setFlags(it.flags() | QtCore.Qt.ItemIsEditable)
                self.setItem(r-min_row, c-min_col, it)

            # detect weekdays row by scanning sequences of 7 cells
            row_texts = [self.item(r-min_row, c-min_col).text() for c in range(min_col, max_col+1)]
            for cstart in range(0, len(row_texts)-6):
                if row_texts[cstart:cstart+7] == wk_names:
                    numbers_row = (r-1)-min_row
                    weekdays_row = r-min_row
                    content_start_row = (r+1)-min_row
                    if self.day_cols==(None,None):
                        self.day_cols=(cstart, cstart+7)
                    self.week_blocks.append((numbers_row, weekdays_row, content_start_row))
                    break

        # merges
        for rng in ws.merged_cells.ranges:
            self.setSpan(rng.min_row-min_row, rng.min_col-min_col, rng.max_row-rng.min_row+1, rng.max_col-rng.min_col+1)

        self.update_layout()

    def update_layout(self):
        total_w=sum(self.base_col_widths) or 1; total_h=sum(self.base_row_heights) or 1
        vw=max(1, self.viewport().width()-2); vh=max(1, self.viewport().height()-2)
        sx=vw/total_w; sy=vh/total_h
        for i,w in enumerate(self.base_col_widths): self.setColumnWidth(i, max(20,int(round(w*sx))))
        for i,h in enumerate(self.base_row_heights): self.setRowHeight(i, max(16,int(round(h*sy))))
    def resizeEvent(self, e): super().resizeEvent(e); self.update_layout()

    # ---------- Month numbers logic ----------
    def apply_month_numbers(self):
        """Заполнить строки чисел дней для self.year/self.month. Хвост следующего/предыдущего месяца подсветить серым и сделать нередактируемым."""
        if not self.week_blocks or self.day_cols==(None,None): return
        y, m = self.year, self.month
        first_weekday, days_in_month = calendar.monthrange(y, m)  # Monday=0
        # Build 5 weeks x 7 days matrix of ints and a boolean is_current flag
        weeks = [[{"num":"", "active":False} for _ in range(7)] for _ in range(len(self.week_blocks))]

        # starting index in first week
        idx = first_weekday  # 0..6
        d = 1
        w = 0
        # fill current month
        while d <= days_in_month and w < len(weeks):
            while idx < 7 and d <= days_in_month:
                weeks[w][idx] = {"num":str(d), "active":True}
                d += 1; idx += 1
            w += 1; idx = 0

        # prev month tail for first week if needed
        if self.week_blocks:
            prev_y, prev_m = (y-1,12) if m==1 else (y,m-1)
            _, prev_days = calendar.monthrange(prev_y, prev_m)
            first_w = weeks[0]
            for i in range(7):
                if not first_w[i]["active"]:
                    # fill from right to left
                    count = sum(1 for j in range(7) if not first_w[j]["active"])
            # Fill leading blanks with tail of prev month
            lead = 0
            for i in range(7):
                if not first_w[i]["active"]:
                    lead += 1
            val = prev_days - lead + 1
            for i in range(lead):
                first_w[i] = {"num":str(val+i), "active":False}

        # next month tail in last week(s)
        next_y, next_m = (y+1,1) if m==12 else (y,m+1)
        nd = 1
        for wi in range(len(weeks)-1, -1, -1):
            row = weeks[wi]
            for i in range(7):
                if row[i]["num"]=="":
                    row[i] = {"num":str(nd), "active":False}; nd += 1

        # apply to table
        start_col, end_col = self.day_cols
        inactive_brush = QtGui.QBrush(QtGui.QColor(140,140,140))
        active_brush = QtGui.QBrush()
        for widx, (num_row, _, content_row) in enumerate(self.week_blocks):
            for i in range(7):
                col = start_col + i
                it = self.item(num_row, col) or QtWidgets.QTableWidgetItem()
                it.setText(weeks[widx][i]["num"])
                # grey inactive
                if weeks[widx][i]["active"]:
                    it.setForeground(active_brush); it.setFlags((it.flags() | QtCore.Qt.ItemIsEditable))
                else:
                    it.setForeground(inactive_brush); it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.setItem(num_row, col, it)

                # optionally grey out the entire content block below for inactive days
                # Здесь мы просто делаем первую строку контента нередактируемой/серой для неактивных дней
                content_it = self.item(content_row, col)
                if content_it:
                    if weeks[widx][i]["active"]:
                        content_it.setForeground(active_brush); content_it.setFlags(content_it.flags() | QtCore.Qt.ItemIsEditable)
                    else:
                        content_it.setForeground(inactive_brush); content_it.setFlags(content_it.flags() & ~QtCore.Qt.ItemIsEditable)

    # ---------- Navigation ----------
    def go_prev_month(self):
        self.save_current_month()
        if self.month==1: self.month=12; self.year-=1
        else: self.month-=1
        # load saved data if exists, else keep current content (Excel baseline)
        if not self.load_month_data(self.year, self.month):
            pass
        self.apply_month_numbers()

    def go_next_month(self):
        self.save_current_month()
        if self.month==12: self.month=1; self.year+=1
        else: self.month+=1
        if not self.load_month_data(self.year, self.month):
            pass
        self.apply_month_numbers()


class CollapsibleSidebar(QtWidgets.QFrame):
    toggled = QtCore.Signal(bool)
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("Sidebar")
        self.setStyleSheet("""
            #Sidebar { background-color: #1f1f23; }
            QToolButton { color: white; border: none; padding: 10px; border-radius: 8px; }
            QToolButton:hover { background-color: rgba(255,255,255,0.08); }
            QLabel { color: #c7c7c7; }
        """)
        self.expanded_width=260; self.collapsed_width=64
        lay=QtWidgets.QVBoxLayout(self); lay.setContentsMargins(8,8,8,8); lay.setSpacing(6)

        # Toggle button — только иконка
        self.btn_toggle = QtWidgets.QToolButton(self)
        self.btn_toggle.setIcon(QtGui.QIcon(ICON_TOGGLE))
        self.btn_toggle.setIconSize(QtCore.QSize(28,28))
        self.btn_toggle.setToolButtonStyle(QtCore.Qt.ToolButtonIconOnly)
        self.btn_toggle.setCursor(QtCore.Qt.PointingHandCursor)
        self.btn_toggle.clicked.connect(self.toggle)
        lay.addWidget(self.btn_toggle)

        line = QtWidgets.QFrame(self); line.setFrameShape(QtWidgets.QFrame.HLine); line.setStyleSheet("color:#333;")
        lay.addWidget(line)

        items = [
            ("Выкладка", ICON_VYKL),
            ("Статистика", ICON_TG),
            ("Топ месяца", ICON_TM),
            ("Топ квартала", ICON_TQ),
            ("Топ полугода", ICON_TP),
            ("Топ года", ICON_TG),
        ]
        self.buttons=[]
        self.btn_release=None
        self.btn_stats=None
        self.btn_top_month=None
        self.btn_top_quarter=None
        self.btn_top_half=None
        self.btn_top_year=None
        for label, icon in items:
            b=QtWidgets.QToolButton(self)
            b.setIcon(QtGui.QIcon(icon)); b.setIconSize(QtCore.QSize(22,22))
            b.setText(" "+label)
            b.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
            b.setCursor(QtCore.Qt.PointingHandCursor)
            b.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
            lay.addWidget(b); self.buttons.append(b)
            if label == "Выкладка":
                self.btn_release = b
            elif label == "Статистика":
                self.btn_stats = b
            elif label == "Топ месяца":
                self.btn_top_month = b
            elif label == "Топ квартала":
                self.btn_top_quarter = b
            elif label == "Топ полугода":
                self.btn_top_half = b
            elif label == "Топ года":
                self.btn_top_year = b

        # --- stats table ---
        self.current_year = datetime.now().year
        self.current_month = datetime.now().month
        self.sorted_col = 0
        self.sorted_order = QtCore.Qt.AscendingOrder

        self.stats_table = QtWidgets.QTableWidget(0, 4, self)
        self.stats_table.setHorizontalHeaderLabels(["Год", "Работа", "Статус", "18+"])
        self.stats_table.horizontalHeader().setStretchLastSection(True)
        self.stats_table.setSortingEnabled(True)
        self.stats_table.horizontalHeader().sortIndicatorChanged.connect(self._update_sort)
        lay.addWidget(self.stats_table)

        # input fields below table
        form = QtWidgets.QWidget(self)
        form_lay = QtWidgets.QFormLayout(form)
        self.edit_work = QtWidgets.QLineEdit(form)
        self.edit_status = QtWidgets.QLineEdit(form)
        self.edit_adult = QtWidgets.QCheckBox("", form)
        form_lay.addRow("Работа", self.edit_work)
        form_lay.addRow("Статус", self.edit_status)
        form_lay.addRow("18+", self.edit_adult)
        self.btn_save_stats = QtWidgets.QPushButton("Сохранить", form)
        self.btn_save_stats.clicked.connect(self.save_stats)
        form_lay.addRow(self.btn_save_stats)
        lay.addWidget(form)
        self.form_widget = form

        lay.addStretch(1)
        self._collapsed=False
        self.anim = QtCore.QPropertyAnimation(self, b"maximumWidth", self); self.anim.setDuration(160)
        self.setMaximumWidth(self.expanded_width)

    def _update_sort(self, col, order):
        self.sorted_col = col
        self.sorted_order = order

    def _add_row(self, work, status, adult):
        row = self.stats_table.rowCount()
        self.stats_table.insertRow(row)
        self.stats_table.setItem(row, 0, QtWidgets.QTableWidgetItem(str(self.current_year)))
        self.stats_table.setItem(row, 1, QtWidgets.QTableWidgetItem(work))
        self.stats_table.setItem(row, 2, QtWidgets.QTableWidgetItem(status))
        adult_item = QtWidgets.QTableWidgetItem("Да" if adult else "Нет")
        adult_item.setData(QtCore.Qt.UserRole, adult)
        self.stats_table.setItem(row, 3, adult_item)

    def _load_stats(self):
        path = os.path.join(STATS_DIR, f"{self.current_year}.json")
        self.stats_table.setRowCount(0)
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            month_data = data.get(str(self.current_month), [])
            for rec in month_data:
                self._add_row(rec.get("work", ""), rec.get("status", ""), rec.get("adult", False))
        if self.stats_table.rowCount() and self.sorted_col is not None:
            self.stats_table.sortItems(self.sorted_col, self.sorted_order)

    def _write_stats_file(self):
        os.makedirs(STATS_DIR, exist_ok=True)
        path = os.path.join(STATS_DIR, f"{self.current_year}.json")
        data = {}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        month_key = str(self.current_month)
        data[month_key] = []
        for row in range(self.stats_table.rowCount()):
            data[month_key].append({
                "work": self.stats_table.item(row,1).text(),
                "status": self.stats_table.item(row,2).text(),
                "adult": bool(self.stats_table.item(row,3).data(QtCore.Qt.UserRole))
            })
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def save_stats(self):
        work = self.edit_work.text().strip()
        status = self.edit_status.text().strip()
        adult = self.edit_adult.isChecked()
        if not work:
            return
        self._add_row(work, status, adult)
        self.stats_table.sortItems(self.sorted_col, self.sorted_order)
        self.edit_work.clear(); self.edit_status.clear(); self.edit_adult.setChecked(False)
        self._write_stats_file()

    def set_period(self, year, month):
        self.current_year = year
        self.current_month = month
        self._load_stats()

    def set_collapsed(self, collapsed: bool):
        self._collapsed=collapsed
        self.anim.stop(); self.anim.setStartValue(self.maximumWidth()); self.anim.setEndValue(self.collapsed_width if collapsed else self.expanded_width); self.anim.start()
        for b in self.buttons:
            b.setToolButtonStyle(QtCore.Qt.ToolButtonIconOnly if collapsed else QtCore.Qt.ToolButtonTextBesideIcon)
        self.stats_table.setVisible(not collapsed)
        self.form_widget.setVisible(not collapsed)
        self.toggled.emit(not collapsed)

    def toggle(self): self.set_collapsed(not self._collapsed)


class SettingsDialog(QtWidgets.QDialog):
    """Окно настроек приложения."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки")
        self.resize(400, 200)
        form = QtWidgets.QFormLayout(self)

        self.chk_neon = QtWidgets.QCheckBox(self)
        self.chk_neon.setChecked(CONFIG.get("neon", False))
        form.addRow("Неоновая подсветка", self.chk_neon)

        self.font_header = QtWidgets.QFontComboBox(self)
        self.font_header.setCurrentFont(QtGui.QFont(CONFIG.get("header_font", "Arial")))
        form.addRow("Шрифт заголовков", self.font_header)

        self.font_text = QtWidgets.QFontComboBox(self)
        self.font_text.setCurrentFont(QtGui.QFont(CONFIG.get("text_font", "Arial")))
        form.addRow("Шрифт текста", self.font_text)

        path_lay = QtWidgets.QHBoxLayout()
        self.edit_path = QtWidgets.QLineEdit(CONFIG.get("save_path", DATA_DIR), self)
        btn_browse = QtWidgets.QPushButton("...", self)
        btn_browse.clicked.connect(self.browse_path)
        path_lay.addWidget(self.edit_path, 1)
        path_lay.addWidget(btn_browse)
        form.addRow("Путь сохранения", path_lay)

        box = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel, self
        )
        box.accepted.connect(self.accept)
        box.rejected.connect(self.reject)
        form.addRow(box)

    def browse_path(self):
        path = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Выбрать папку", self.edit_path.text()
        )
        if path:
            self.edit_path.setText(path)

    def accept(self):
        config = {
            "neon": self.chk_neon.isChecked(),
            "header_font": self.font_header.currentFont().family(),
            "text_font": self.font_text.currentFont().family(),
            "save_path": self.edit_path.text().strip() or DATA_DIR,
        }
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        super().accept()


class TopBar(QtWidgets.QWidget):
    prev_clicked = QtCore.Signal()
    next_clicked = QtCore.Signal()
    settings_clicked = QtCore.Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QtWidgets.QHBoxLayout(self)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(8)
        self.btn_prev = QtWidgets.QToolButton(self)
        self.btn_prev.setText(" < ")
        self.btn_prev.clicked.connect(self.prev_clicked)
        lay.addWidget(self.btn_prev)
        self.lbl_month = QtWidgets.QLabel("Месяц")
        self.lbl_month.setAlignment(QtCore.Qt.AlignCenter)
        lay.addWidget(self.lbl_month, 1)
        self.btn_next = QtWidgets.QToolButton(self)
        self.btn_next.setText(" > ")
        self.btn_next.clicked.connect(self.next_clicked)
        lay.addWidget(self.btn_next)
        self.btn_settings = QtWidgets.QToolButton(self)
        self.btn_settings.setText("⚙")
        self.btn_settings.clicked.connect(self.settings_clicked)
        lay.addWidget(self.btn_settings)
        self.default_style = (
            "QLabel{color:#e5e5e5;} QToolButton{color:#e5e5e5; border:1px solid #555; border-radius:6px; padding:4px 10px;}"
        )
        self.neon_style = (
            "QLabel{color:#39ff14;} QToolButton{color:#39ff14; border:1px solid #39ff14; border-radius:6px; padding:4px 10px;}"
        )
        pal = self.palette()
        pal.setColor(self.backgroundRole(), QtGui.QColor(30, 30, 33))
        self.setAutoFillBackground(True)
        self.setPalette(pal)
        self.apply_style(CONFIG.get("neon", False))
        self.apply_fonts()

    def apply_fonts(self):
        font = QtGui.QFont(CONFIG.get("header_font", "Arial"))
        font.setBold(True)
        self.lbl_month.setFont(font)

    def apply_style(self, neon):
        self.setStyleSheet(self.neon_style if neon else self.default_style)


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("План-график — Excel 1:1 + навигация по месяцам + левый бар")
        central = QtWidgets.QWidget(self)
        h = QtWidgets.QHBoxLayout(central); h.setContentsMargins(0,0,0,0); h.setSpacing(0)

        # left sidebar
        self.sidebar=CollapsibleSidebar(self); h.addWidget(self.sidebar)

        # right: vbox with topbar + table
        right = QtWidgets.QWidget(self); v = QtWidgets.QVBoxLayout(right); v.setContentsMargins(0,0,0,0); v.setSpacing(0)
        self.topbar = TopBar(self); v.addWidget(self.topbar)
        self.table = ExcelCalendarTable(self); v.addWidget(self.table, 1)
        h.addWidget(right, 1)

        self.setCentralWidget(central)
        self.showMaximized()

        # Connect topbar
        self.topbar.prev_clicked.connect(self.prev_month)
        self.topbar.next_clicked.connect(self.next_month)
        self.sidebar.btn_release.clicked.connect(self.open_release_dialog)
        self.sidebar.btn_stats.clicked.connect(self.open_year_stats_dialog)
        self.sidebar.btn_top_month.clicked.connect(lambda: self.open_top_dialog("month"))
        self.sidebar.btn_top_quarter.clicked.connect(lambda: self.open_top_dialog("quarter"))
        self.sidebar.btn_top_half.clicked.connect(lambda: self.open_top_dialog("half"))
        self.sidebar.btn_top_year.clicked.connect(lambda: self.open_top_dialog("year"))
        self.topbar.settings_clicked.connect(self.open_settings_dialog)
        self._update_month_label()
        self.sidebar.set_period(self.table.year, self.table.month)

    def _update_month_label(self):
        self.topbar.lbl_month.setText(f"{RU_MONTHS[self.table.month-1]} {self.table.year}")

    def prev_month(self):
        self.table.go_prev_month(); self._update_month_label()
        self.sidebar.set_period(self.table.year, self.table.month)

    def next_month(self):
        self.table.go_next_month(); self._update_month_label()
        self.sidebar.set_period(self.table.year, self.table.month)

    def open_release_dialog(self):
        dlg = ReleaseDialog(self.table.year, self.table.month, self)
        dlg.exec()

    def open_year_stats_dialog(self):
        dlg = YearStatsDialog(self.table.year, self)
        dlg.exec()

    def open_top_dialog(self, mode):
        if mode == "month":
            period = self.table.month
        elif mode == "quarter":
            period = (self.table.month - 1) // 3 + 1
        elif mode == "half":
            period = (self.table.month - 1) // 6 + 1
        else:
            period = 1
        dlg = TopDialog(self.table.year, mode, period, self)
        dlg.exec()

    def open_settings_dialog(self):
        dlg = SettingsDialog(self)
        if dlg.exec():
            global CONFIG, BASE_SAVE_PATH, STATS_DIR, YEAR_DIR, RELEASE_DIR, TOP_DIR
            CONFIG = load_config()
            BASE_SAVE_PATH = os.path.abspath(CONFIG.get("save_path", DATA_DIR))
            STATS_DIR = os.path.join(BASE_SAVE_PATH, "stats")
            YEAR_DIR = os.path.join(BASE_SAVE_PATH, "year")
            RELEASE_DIR = os.path.join(BASE_SAVE_PATH, "release")
            TOP_DIR = os.path.join(BASE_SAVE_PATH, "top")
            app = QtWidgets.QApplication.instance()
            app.setFont(QtGui.QFont(CONFIG.get("text_font", "Arial")))
            self.apply_settings()

    def apply_settings(self):
        self.topbar.apply_fonts()
        self.topbar.apply_style(CONFIG.get("neon", False))


def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setFont(QtGui.QFont(CONFIG.get("text_font", "Arial")))
    w = MainWindow()
    w.apply_settings()
    w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
